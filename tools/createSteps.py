# -*- coding = utf-8 -*-
# @Time: 2023/7/13 20:49
# @Author: Gerry
# @File: createSteps.py
# @Software: PyCharm
from globalImportBase import zPrint, Excel, re, os, changePath
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
filePath = "D:/userData/自动化不可实现步骤/"

def createSteps():
    """
    创建步骤
    :return:
    """
    zPrint("开始创建步骤")
    Excel.genCaseStep(excelPath="D:\\test.xlsx", sheet_name="Sheet1", purColNameList=["采购方"], supColNameList=["供应商"])
    zPrint("步骤创建完成")

def readUnCreateStep(dirStr="."):
    """
    读取步骤中的不可实现步骤
    :return:
    """
    lineData = []
    for dirpath, dirnames, filenames in os.walk(dirStr):
        for filename in filenames:
            filePathStr = os.path.join(dirpath, filename)
            # 获取用例文件, 去除缓存文件
            if "testXT" in filePathStr and "__pycache__" not in filePathStr:
                zPrint("文件路径",filePathStr)
                # 用例名称
                caseCode = os.path.basename(filePathStr).split(".")[0].replace("test", "")
                zPrint("用例名称",caseCode)
                # 打开文件
                fo = open(filePathStr, "r", encoding="utf8")
                content = fo.read(-1)

                # 获取负责人
                pattern = r"@Author: (.*)"
                authorList = re.findall(pattern, content)
                # assert len(authorList) == 1, f"用例{caseCode}负责人输入有误"
                author = authorList[0] if len(authorList) == 1 else f"负责人编写有误"
                zPrint("负责人", author)

                # 场景
                pattern = "@allure.story\(\"(.*)\"\)"
                storyList = re.findall(pattern, content)
                # assert len(authorList) == 1, f"用例{caseCode}场景输入有误"
                story = storyList[0]  if len(authorList) == 1 else f"场景编写有误"
                zPrint(story)

                # 通过正则表达式获取到不可实现的步骤
                pattern = r"&(.*)&"
                stepList = re.findall(pattern, content)
                zPrint("不可实现步骤", stepList)
                # lines = []
                for step in stepList:
                    line = []
                    steps = str(step).split("|")
                    # assert len(steps) == 2, f"用例{caseCode}, {stepList} 有误"
                    steps = steps if len(steps) == 2 else [f"用例{caseCode}不可实现步骤的标记有误",f"{steps}"]
                    line.append(caseCode)
                    line.append(story)
                    line.extend(steps)
                    line.append(author)
                    # lines.append(line)
                    lineData.append(line)
                else:
                    lineData.append([caseCode,story,"","", author])
                # 关闭文件
                fo.close()

    # 写入文件
    writeExcel(lineData=lineData)

# 写入Excel
def writeExcel(lineData):
    """
    循环查询写入Excel
    :return:
    """
    fileName = "{}自动化不可实现步骤.xlsx".format(filePath)
    if os.path.exists(filePath) is True:
        # 有Excel打开, 无Excel文件创建新的Excel,
        if os.path.exists(fileName) is True:
            wb = load_workbook(fileName)
        else:
            # 实例化
            wb = Workbook()
    else:
        # 创建目录
        os.makedirs(filePath)
        # 实例化
        wb = Workbook()

    # 创建sheet页
    sheet = wb.create_sheet(index=0, title="协同不可实现步骤")  # 在第一个位置插入工作表
    headerList = ["用例编号","场景", "不可实现步骤编号", "不可实现原因", "脚本负责人"]
    sheet.append(headerList)
    for line in lineData:
        sheet.append(line)

    # wb 属性设置
    # 设置行高和列宽
    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = 20

    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 30


    # 对齐方式
    alight = Alignment(
        horizontal='left',  # 水平对齐方式:center, left, right
        vertical='center'  # 垂直对齐方式: center, top, bottom
    )
    font = Font(
        size=12,
        italic=False,
        color='000000',
        bold=True,
        strike=None
    )
    bold_style = Side(border_style='thin', color='000000')
    border = Border(left=bold_style, right=bold_style, top=bold_style, bottom=bold_style)
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            sheet.cell(row=i, column=j).border = border
            sheet.cell(row=i, column=j).alignment = alight

    blue_fill = PatternFill(fgColor="87CEEB", fill_type="solid")
    for i in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=i).fill = blue_fill
        sheet.cell(row=1, column=i).font = font


    # 保存Excel到指定目录
    wb.save(fileName)

if __name__ == '__main__':
    createSteps()
    # casePath = changePath("SrmInterfaceTest", "/case/standard/choiceForm")
    # readUnCreateStep(casePath)