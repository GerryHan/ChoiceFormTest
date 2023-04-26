import datetime
import random
import string
import time
import os
import re
import requests
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
from decimal import Decimal
import calendar


def alter(file, old_str, new_str):
    """
    替换文件中的字符串
    :param file:文件名
    :param old_str:老字符串
    :param new_str:新字符串
    :return:
    """
    file_data = ""
    with open(file, "r", encoding="utf-8") as f:
        for line in f:
            if old_str in line:
                line = line.replace(old_str, new_str)
            file_data += line
    with open(file, "w", encoding="utf-8") as f:
        f.write(file_data)


def randomNumNotRepeat(start=0, end=100, n=1):
    """
    创建n个0~num的随机整数
    :param end: 开始
    :param start: 结束
    :param n: 随机数的个数
    :return:
    """
    if start >= end:
        return 0
    if end - start < n:
        return "随机个数超出随机范围"
    return random.sample(range(start, end), n)


priorityList = []


def randomPriority(start=1, end=1000, n=1):
    """
    创建n个0~num的随机整数
    :param end: 开始
    :param start: 结束
    :param n: 随机数的个数
    :return:
    """
    if start >= end:
        return 0
    if end - start < n:
        return "随机个数超出随机范围"
    priority = random.sample(range(start, end), n)[0]
    priorityList.append(priority)
    for i in range(100):
        if priority in priorityList:
            priority = random.sample(range(start, end), n)[0]
        else:
            return priority
    else:
        assert False, "生成的优先级重复了"


def random_string(str_len, letter_digits="DandL"):
    """
    生成随机字符串
    :param letter_digits: 随机数类型 D：数字 L：字母
    :param str_len 字符串长度
    """
    if letter_digits.upper() == "D":
        return "".join(random.choice(string.digits) for _ in range(str_len))
    elif letter_digits.upper() == "L":
        return "".join(random.choice(string.ascii_letters) for _ in range(str_len))
    else:
        return "".join(random.choice(string.ascii_letters + string.digits) for _ in range(str_len))


def randomShuffle(data, fixValue=None, fixLen=10, upper=False):
    """
    随机生成字符串并打乱顺序
    data={
        "digit":len,#数字类型，value值代表随机生成的长度
        "letter":len,#字符串类型，value值代表随机生成的长度
    }
    fixValue:固定字符串长度
    upper:字母是否大写
    """
    randomStr = ""
    if data.get("digit") and data.get("letter") is None:
        randomStr += "".join(random.choice(string.digits) for _ in range(data.get("digit")))
    elif data.get("letter") and data.get("digit") is None:
        if upper is True:
            randomStr += "".join(random.choice(string.ascii_letters) for _ in range(data.get("letter"))).upper()
        else:
            randomStr += "".join(random.choice(string.ascii_letters) for _ in range(data.get("letter")))
    elif data.get("letter") and data.get("digit"):
        if upper is True:
            randomStr += "".join(random.choice(string.ascii_letters) for _ in range(data.get("letter"))).upper()
        else:
            randomStr += "".join(random.choice(string.ascii_letters) for _ in range(data.get("letter")))
        randomStr += "".join(random.choice(string.digits) for _ in range(data.get("digit")))
    else:
        randomStr += "".join(random.choice(string.ascii_letters + string.digits) for _ in range(fixLen))
    randomList = list(randomStr)
    random.shuffle(randomList)
    randomStr = ''.join(randomList)
    if fixValue:
        randomStr = str(fixValue) + randomStr
    return randomStr


def timestamp(str_len=13):
    """
    生成时间戳
    :param str_len 截取长度
    """
    if isinstance(str_len, int) and 0 < str_len < 17:
        return str(time.time()).replace(".", "")[:str_len]


def days(str1, str2, hours=None, minute=None, seconds=None):
    """
    计算日期间差值
    @param str1:
    @param str2:
    @return:
    """
    # 计算秒数之差
    if seconds:
        startTime = datetime.datetime.strptime(str1, "%Y-%m-%d %H:%M:%S")
        endTime = datetime.datetime.strptime(str2, "%Y-%m-%d %H:%M:%S")
        # 来获取时间差中的秒数。注意，seconds获得的秒只是时间差中的小时、分钟和秒部分的和，并没有包含时间差的天数（既是两个时间点不是同一天，失效）
        secondNums = (endTime - startTime).total_seconds()
        # 来获取准确的时间差，并将时间差转换为秒
        return int(secondNums)
    # 计算分钟数之差
    if minute:
        startTime = datetime.datetime.strptime(str1, "%Y-%m-%d %H:%M:%S")
        endTime = datetime.datetime.strptime(str2, "%Y-%m-%d %H:%M:%S")
        # 来获取时间差中的秒数。注意，seconds获得的秒只是时间差中的小时、分钟和秒部分的和，并没有包含时间差的天数（既是两个时间点不是同一天，失效）
        secondNums = (endTime - startTime).total_seconds()
        # 来获取准确的时间差，并将时间差转换为秒
        return int(secondNums / 60)
    # 计算小时数之差
    if hours:
        startTime = datetime.datetime.strptime(str1, "%Y-%m-%d %H:%M:%S")
        endTime = datetime.datetime.strptime(str2, "%Y-%m-%d %H:%M:%S")
        # 来获取时间差中的秒数。注意，seconds获得的秒只是时间差中的小时、分钟和秒部分的和，并没有包含时间差的天数（既是两个时间点不是同一天，失效）
        secondNums = (endTime - startTime).total_seconds()
        # 来获取准确的时间差，并将时间差转换为秒
        return int(secondNums / 3600)
    else:
        date1 = datetime.datetime.strptime(str1[0:10], "%Y-%m-%d")
        date2 = datetime.datetime.strptime(str2[0:10], "%Y-%m-%d")
        num = (date1 - date2).days
        return num


def date_time(timezone=None, years=0, months=0, days=0, hours=0, minutes=0, seconds=0, microseconds=0, fmt="%Y-%m-%d"):
    # 拓展支持年月计算 2023-01-10 伍超超
    """
    获取当前时间,或者在当前时间增加或者减少相应的时间,
    :param years: 增加或减少的年数, 正数为加,负数为减
    :param months: 增加或减少的月数, 正数为加,负数为减
    :param days: 增加或减少的天数, 正数为加,负数为减
    :param hours: 增加或减少的小时数
    :param minutes: 增加或减少的分钟数
    :param seconds: 增加或减少的秒钟数
    :param microseconds: 增加或减少的微秒钟数
    :param fmt: 输出格式,默认为 %Y-%m-%d,如:2020-06-02, 可指定其他格式  %Y-%m-%d %H:%M:%S
    :param timezone: 时区
    :return:
    """
    if timezone:
        now = eval(f"datetime.datetime.now(tz=datetime.timezone.{timezone})")
    else:
        now = datetime.datetime.now()
    result_date = now + relativedelta(
        years=years,
        months=months,
        days=days,
        hours=hours,
        minutes=minutes,
        seconds=seconds,
        microseconds=microseconds,
    )
    return result_date.strftime(fmt)


def strToDate(dateStr, days=0, hours=0, minutes=0, seconds=0, inFmt="%Y-%m-%d", outFmt="%Y-%m-%d"):
    """
    获取当前时间,或者在当前时间增加或者减少相应的时间,
    :param dateStr: 日期字符串
    :param inFmt: 输入格式,默认为 %Y-%m-%d,如:2023-06-15, 可指定其他格式%Y-%m-%d %H:%M:%S
    :param days: 增加或减少的天数, 正数为加,负数为减
    :param hours: 增加或减少的小时数
    :param minutes: 增加或减少的分钟数
    :param outFmt: 输出格式,默认为 %Y-%m-%d,如:2020-06-02, 可指定其他格式%Y-%m-%d %H:%M:%S
    :return:
    """
    date = datetime.datetime.strptime(dateStr, inFmt)
    return (date + datetime.timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds)).strftime(outFmt)


def getYearFirstDayAndLastDay(year=None, fmt="%Y-%m-%d"):
    """
    获取某年的第一天，最后一天
    :param year:
    :param fmt:
    :return:
    """
    if year is None:
        # 获取当前年份
        year = datetime.date.today().year
    # 获取某年的第一天
    firstDay = datetime.date(year, 1, day=1)
    # 获取某年的最后一天
    lastDay = datetime.date(year, 12, day=31)
    return firstDay.strftime(fmt), lastDay.strftime(fmt)


def getMonthFirstDayAndLastDay(year=None, month=None, fmt="%Y-%m-%d"):
    """
    获取某年某月的第一天，最后一天
    :param year:
    :param month:
    :param fmt:
    :return:
    """
    if year is None:
        # 获取当前年份
        year = datetime.date.today().year
    if month is None:
        # 获取当前月份
        month = datetime.date.today().month
    # 获取当前月的第一天的星期和当月总天数
    weekDay, monthCountDay = calendar.monthrange(year, month)
    # 获取当前月份第一天
    firstDay = datetime.date(year, month, day=1)
    # 获取当前月份最后一天
    lastDay = datetime.date(year, month, day=monthCountDay)
    return firstDay.strftime(fmt), lastDay.strftime(fmt)


def getCycleDate(fromDay, cycle='MONTH', inFmt="%Y-%m-%d", outFmt="%Y-%m-%d"):
    """
    通过开始日期、周期，获取截至日期
    :param fromDay: 开始日期 2023-01-01
    :param cycle: 周期：MONTH 月度、QUARTER 季度、HALF-YEAR 半年度、 YEAR年度
    :param inFmt: 输入格式,默认为 %Y-%m-%d,如:2023-06-15, 可指定其他格式%Y-%m-%d %H:%M:%S
    :param outFmt: 输出格式,默认为 %Y-%m-%d,如:2020-06-02, 可指定其他格式%Y-%m-%d %H:%M:%S
    :return:
    """
    fromDay = datetime.datetime.strptime(fromDay, inFmt)
    if cycle == 'MONTH' or cycle == '月度':  # 月度
        toDay = fromDay + relativedelta(years=0, months=1, days=-1, hours=0, minutes=0, seconds=0)
    elif cycle == 'QUARTER' or cycle == '季度':  # 季度
        toDay = fromDay + relativedelta(years=0, months=3, days=-1, hours=0, minutes=0, seconds=0)
    elif cycle == 'HALF-YEAR' or cycle == '半年度':  # 半年度
        toDay = fromDay + relativedelta(years=0, months=6, days=-1, hours=0, minutes=0, seconds=0)
    elif cycle == 'YEAR' or cycle == '年度':  # 年度
        toDay = fromDay + relativedelta(years=1, months=0, days=-1, hours=0, minutes=0, seconds=0)
    else:
        return None
    return toDay.strftime(outFmt)


def sleep(sec):
    # 停顿*秒
    time.sleep(sec)


def executeTime(fun):
    """
    计算方法的执行时间
    :param fun:
    :return:
    """

    def wrapper(*args, **kwargs):
        # 定义开始时间和结束时间，将fun夹在中间执行，取得其返回值
        start = datetime.datetime.now()
        func_return = fun(*args, **kwargs)
        end = datetime.datetime.now()
        # 打印方法名称和其执行时间
        print(f'{fun.__name__}() execute time: {end - start}s')
        # 返回func的返回值
        return func_return

    # 返回嵌套的函数
    return wrapper


def changePath(midDir, rightPath):
    """
    更换目录
    :param midDir: 目录分割点 (目录分叉点)
    :param rightPath:
    :return:
    """
    # 获取当前文件路径
    path = os.path.dirname(os.path.abspath(__file__))
    # 从目录分叉点分隔
    pathList = path.rsplit(midDir, 1)
    # 替换并拼接
    fullPath = pathList[0].replace("\\", "/") + midDir + rightPath
    return fullPath


def saveExportFile(response, fileName):
    """
    保存导出文件
    ※※注：导出文件接口需传parseToJson=None，否则无法获取到content。例request.get(url, scene=scene, parseToJson=None)
    :param response: 导出接口response
    :param fileName: 文件名称※※注：文件名称需与其他文件不同，最好起该用例的名称
    :return:
    """
    filePath = os.path.abspath(__file__)[0:3:1]
    if "C:" in filePath:
        path = filePath.replace("C:", "D:")
    else:
        path = filePath
    openFile = open(f'{path}{fileName}', 'wb')
    openFile.write(response.content)
    openFile.close()


def saveUrlExportFile(response, fileName, needParse=True):
    """
    下载返回值为URL的文件
    :param response: 接口不解析的返回值
    :param fileName: 文件名称※※注：文件名称需与其他文件不同，最好起该用例的名称
    :return:
    """
    if needParse:
        # 下载地址
        Download_addres = response.text
    else:
        Download_addres = response
    # 把下载地址发送给requests模块
    f = requests.get(Download_addres)
    # 下载文件
    filePath = os.path.abspath(__file__)[0:3:1]
    if "C:" in filePath:
        path = filePath.replace("C:", "D:")
    else:
        path = filePath
    with open(f'{path}{fileName}', "wb") as code:
        code.write(f.content)
        code.close()
    return path + fileName


def deleteFile(fileName):
    """
    删除文件
    :param fileName: 文件名称※※注：文件名称需与其他文件不同，最好起该用例的名称
    :return:
    """
    filePath = os.path.abspath(__file__)[0:3:1]
    if "C:" in filePath:
        path = filePath.replace("C:", "D:")
    else:
        path = filePath
    dataPath = f'{path}{fileName}'
    if os.path.exists(dataPath):
        os.remove(dataPath)
    else:
        print('no such file')


def confirmUnique(contentList, **kwargs):
    """
    从字典列表中筛选出唯一的字典
    :param contentList: 选择列表
    :param kwargs: 选择条件
    :return:
    """
    assert isinstance(contentList, list), "请传入数组"
    if len(contentList) > 0:
        confirmList = []
        for item in contentList:
            for parm in kwargs.keys():
                # 若有一个不相等,跳出循环查找下一个
                if kwargs.get(parm) is not None and item.get(parm) != kwargs.get(parm):
                    break
            else:
                # 若全部相等, 添加到已经确认的数组中
                confirmList.append(item)
        else:
            # 若唯一返回, 若不唯一,需要更多的字段来筛选
            if len(confirmList) == 1:
                return confirmList[0]
            else:
                assert False, "请使用更多字段筛选"
    else:
        assert False, "请传入正确列表"


def screenDict(contentList, **kwargs):
    """
    从列表中筛选指定的字典
    :param contentList: 选择列表
    :param key: 键
    :return:
    """
    assert isinstance(contentList, list), "请传入数组"
    if len(contentList) > 0:
        confirmList = []
        for item in contentList:
            for key, val in kwargs.items():
                # 若有一个不相等,跳出循环查找下一个
                if item.get(key) == val:
                    confirmList.append(item)
        return confirmList

    else:
        assert False, "请传入正确列表"


def noneToEmptyString(parm):
    if parm is None:
        return ""
    else:
        return parm


def getPrecision(parm):
    """
    获取精度-数字小数部分的长度; 注意:如果数字后面只带小数点 eg 397. 则精度为1
    :param parm:
    :return:
    """
    assert parm is not None, "请传入数字"
    assert type(parm) in (float, int)
    li = str(parm).split(".")
    if len(li) > 1:
        return len(li[1])
    else:
        return 0


def cleanNoneFromDict(parm):
    """
    清除字典中值为None的键值对
    :param parm:
    :return:
    """
    if isinstance(parm, dict):
        return {k: v for k, v in parm.items() if (v is not None)}
    else:
        assert False, "请传入字典"


def sortDictInList(listA, key, reverse=False):
    """
    内建函数 sorted 方法返回的是一个新的 list，而不是在原来的基础上进行的操作。
    :param reverse:
    :param listA:iterable -- 可迭代对象。
    :param key:key -- 主要是用来进行比较的元素，只有一个参数，具体的函数的参数就是取自于可迭代对象中，指定可迭代对象中的一个元素来进行排序。
    :return:reverse -- 排序规则，reverse = True 降序 ， reverse = False 升序（默认）。
    """
    return sorted(listA, key=lambda d: d[key], reverse=reverse)


def sortDictIntInList(listA, key, reverse=False):
    """
    注意: 只针对key对应的值为数字,或者为整数的数字字符串的
    内建函数 sorted 方法返回的是一个新的 list，而不是在原来的基础上进行的操作。
    :param reverse:
    :param listA:iterable -- 可迭代对象。
    :param key:key -- 主要是用来进行比较的元素，只有一个参数，具体的函数的参数就是取自于可迭代对象中，指定可迭代对象中的一个元素来进行排序。
    :return:reverse -- 排序规则，reverse = True 降序 ， reverse = False 升序（默认）。
    """
    return sorted(listA, key=lambda d: int(d[key]), reverse=reverse)


def sortList(data, key, specChar=None):
    """
    单号排序
    用法：
    场景①：返回值单号为CON20230824000201|3类型，key传此单号的键，specChar传"|"
    场景②：返回值单号为REQ2023082600001，但接口返回根据lineNum排序，key传lineNum，specChar不传
    @param data:列表data值
    @param key: 单号的键
    @param specChar: 特殊字符过滤，如CON20230824000201|3，specChar传"|"
    @return:
    """
    n = len(data)
    for i in range(n):
        for j in range(0, n - i - 1):
            if specChar:
                sortAList, sortBList = data[j].get(key).split(specChar), data[j + 1].get(key).split(specChar)
                if sortAList[0] == sortBList[0]:
                    if int(sortAList[1]) > int(sortBList[1]):
                        data[j], data[j + 1] = data[j + 1], data[j]
                else:
                    numA, numB = re.sub("[A-Za-z]", "", sortAList[0]), re.sub("[A-Za-z]", "", sortBList[0])
                    if int(numA) > int(numB):
                        data[j], data[j + 1] = data[j + 1], data[j]
                        sortList(data=data, key=key, specChar=specChar)
            else:
                sortAList, sortBList = data[j].get(key), data[j + 1].get(key)
                if isinstance(sortAList, int) and isinstance(sortBList, int):
                    numA, numB = sortAList, sortBList
                else:
                    numA, numB = re.sub("[A-Za-z]", "", sortAList), re.sub("[A-Za-z]", "", sortBList)
                if int(numA) > int(numB):
                    data[j], data[j + 1] = data[j + 1], data[j]
                    sortList(data=data, key=key, specChar=specChar)
    return data


def sortDict(dictA, key=0, reverse=False):
    """
    传递一个字典，默认以key值并升序排列，返回一个字典
    传递一个列表，默认升序排列，返回一个字典
    内建函数 sorted 方法返回的是一个新的 list，而不是在原来的基础上进行的操作。
    :param reverse:
    :param dictA:iterable -- 可迭代对象。
    :param key:key -- 主要是用来进行比较的元素索引，0：代表着key值作为比较值，1代表着value值作为比较内容
    :return:reverse -- 排序规则，reverse = True 降序 ， reverse = False 升序（默认）。
    """
    if isinstance(dictA, dict):
        return dict(sorted(dictA.items(), key=lambda d: d[key], reverse=reverse))
    elif isinstance(dictA, list) and isinstance(dictA[0], (str, int)):
        return sorted(dictA, reverse=reverse)


def changeNonString(string):
    """
    将非纯数字的字符串进行加1操作
    :param string:字符串
    :return:
    """
    rt = re.search(r'(\d+)([^\d]*$)', string)
    if rt:
        pos_left = rt.span()[0]
        num = int(rt.groups()[0])
        numStr = ''
        for i in range(len(rt.groups()[0]) - len(str(num + 1))):
            numStr += '0'
        else:
            numStr += str(num + 1)
            return string[:pos_left] + numStr[len(numStr) - len(rt.groups()[0]):] + rt.groups()[1]
    else:
        raise ValueError('No suitable number segment found to +1.')


def writeExcelData(rightPath, sheetName, **kwargs):
    """
    向指定excel写入数据
    :param rightPath: 文件路径
    :param sheetName: 要操作sheet页的名字
    :param kwargs: 字典，键：excel坐标，eg：'A1'，‘B2’；值为：需要写入的值
    :return:
    """
    mulFilePath = changePath("SrmInterfaceTest", rightPath=rightPath)
    wb = load_workbook(mulFilePath)
    sheet = wb.get_sheet_by_name(sheetName)
    for key, value in kwargs.items():
        sheet[key] = value
    wb.save(mulFilePath)
    print("写入数据成功")


def writeExcelDataBtn(filePath, sheetName, **kwargs):
    """
    向指定excel写入数据
    :param filePath: 文件路径
    :param sheetName: 要操作sheet页的名字
    :param kwargs: A = [1,2] : 向 A 列写入 A2 和 A3
    :return:
    """
    wb = load_workbook(filePath)
    sheet = wb.get_sheet_by_name(sheetName)
    kws = {}
    for key, value in kwargs.items():
        for i in range(len(value)):
            kws[key + str(i + 2)] = value[i]
    for key, value in kws.items():
        sheet[key] = value
    wb.save(filePath)
    print("写入数据成功")


def modifySpecifiedField(data, dictField={}):
    """
    :param data: 需要修改的字典数据
    :param dictField: 存储需要修改为value的指定的字段
    """
    for key, value in dictField.items():
        if isinstance(key, super):
            keys = key
        else:
            keys = (key)
        for i in keys:
            data[i] = value
    return data


def listDeDuplication(data, reverse=False):
    """
    :param data: 需要去重的列表
    :param reverse: 默认排序方式为升序
    """
    resultDict = {}  # key值为列表中的元素，value值为元素在列表中出现的次数
    list01 = list(set(data))
    for listData in list01:
        resultDict[listData] = data.count(listData)
    resultDict = dict(sorted(resultDict.items(), key=lambda x: x[0].upper(), reverse=reverse))
    return resultDict


def zRound(floatNum, n=2):  # floatNum为原小数，n为要保留的小数位数
    """
    float类型数据 保留n位小数
    :param floatNum:
    :param n:
    :return:
    """
    formatStr = "0." + "0" * n
    decNum = Decimal(str(floatNum)).quantize(Decimal(formatStr), rounding="ROUND_HALF_UP")
    return float(decNum)

def getFileName(filePath):
    """
    根据文件路径列表或路径获取文件名称列表
    :param filePath:
    :return:
    """
    # 根据路径获取文件名成
    nameList = []
    if isinstance(filePath, list):
        for name in filePath:
            pathList = name.split("/")
            assert len(pathList) > 0
            nameList.append(pathList[-1])
    elif isinstance(filePath, str):
        pathList = filePath.split("/")
        assert len(pathList) > 0
        nameList.append(pathList[-1])
    else:
        assert False, "传入数据格式错误"
    return nameList

def calUnitPriceWithTaxIncludedPrice(taxIncludedPrice, taxRate=0.13, unitPriceBatch=1, quantity=1, precision=None):
    """
    基准价=含税：
    不含税单价=不含税金额/数量*每
    含税金额=含税单价*数量/每
    不含税金额=含税金额-税额；税额=含税金额/（1+税率）*税率
    :param taxIncludedPrice: 含税单价
    :param taxRate: 税率
    :param quantity: 数量 可以省去
    :param unitPriceBatch: 每
    :param precision: 精度
    :return: 不含税单价
    """
    assert quantity > 0
    assert unitPriceBatch > 0

    # 含税金额
    amt = taxIncludedPrice * quantity / unitPriceBatch

    # 税额
    taxAmt = amt / (1 + taxRate) * taxRate

    # 不含税金额
    unitAmt = amt - taxAmt

    # 不含税单价
    unitPrice = unitAmt / quantity * unitPriceBatch

    if precision is not None:
        unitPrice = round(unitPrice, precision)

    return unitPrice


def calTaxIncludedPriceWithUnitPrice(unitPrice, taxRate=0.13, unitPriceBatch=1, quantity=1, precision=None):
    """
    基准价=不含税：
    含税单价=含税金额/数量*每
    不含税金额=不含税单价*数量/每
    含税金额=不含税金额+税额；税额=不含税金额*税率
    :param unitPrice: 不含税单价
    :param taxRate: 税率
    :param quantity: 数量
    :param unitPriceBatch: 每
    :param precision: 精度
    :return: 含税单价
    """
    assert quantity > 0
    assert unitPriceBatch > 0

    # 不含税金额
    unitAmt = unitPrice * quantity / unitPriceBatch

    # 税额
    taxAmt = unitAmt * taxRate

    # 含税金额
    taxIncludeAmt = unitAmt + taxAmt

    # 不含税单价
    taxIncludedPrice = taxIncludeAmt / quantity * unitPriceBatch

    if precision is not None:
        taxIncludedPrice = round(taxIncludedPrice, precision)

    return taxIncludedPrice


def dataReplace(data, info):
    """
    自己替换数据
    @param data: 自己数据
    @param info: 查询出来的数据
    @return:
    """

    for dataKey, dataVlaue in data.items():
        if isinstance(dataVlaue, list) and dataVlaue != []:
            dataVlaueList = [i for i in dataVlaue if i.get("add") == "add"]
            if len(dataVlaueList) == 0:
                for i in range(len(dataVlaue)):
                    info[dataKey][i].update(dataVlaue[i])
            else:
                for i in range(len(dataVlaue) - len(dataVlaueList)):
                    info[dataKey][i].update(dataVlaue[i])
                for i in range(len(dataVlaueList)):
                    del dataVlaueList[i]["add"]
                    info[dataKey].insert(i, dataVlaueList[i])
        else:
            info[dataKey] = dataVlaue


def dataReplaceOne(data, info):
    """
    自己替换数据
    @param data: 自己数据
    @param info: 查询出来的数据
    @return:
    """
    for dataKey, dataVlaue in data.items():
        if isinstance(dataVlaue, list):
            lenN = len(info[dataKey])
            n = len(dataVlaue)
            for i in range(n):
                if len(dataVlaue) > lenN:
                    info[dataKey].append(dataVlaue[i])
                    info[dataKey][i].update(dataVlaue[i])
                    lenN = len(info[dataKey])
                elif len(dataVlaue) == lenN:
                    info[dataKey][i].update(dataVlaue[i])
                elif len(dataVlaue) < lenN:
                    for d in range(len(info[dataKey]) - len(dataVlaue)):
                        info[dataKey].pop()
        else:
            info[dataKey] = dataVlaue


def dataReplaceEase(data, info):
    """
    自己替换数据
    @param data: 自己数据
    @param info: 查询出来的数据
    @return:
    """
    for rep in info:
        for d in data:
            rep.update(d)


def replaceData(myData, standData):
    """
    数据替换,以dict形式传入
    @param myData: 需要修改的数据
    @param standData: 标准数据
    @return:
    """
    if isinstance(myData, dict):
        for key, value in myData.items():
            if myData[key]:
                if isinstance(value, list) or isinstance(value, dict):
                    replaceData(myData[key], standData[key])
                    if isinstance(value, list):
                        if type(value[0]) is not dict:
                            standData[key] = myData.get(key)
                        else:
                            for myD in myData[key]:
                                standData[key][myData[key].index(myD)].update(myD)
                else:
                    standData[key] = myData.get(key)
    elif isinstance(myData, list):
        for i in myData:
            if isinstance(i, list) or isinstance(i, dict):
                replaceData(i, standData[myData.index(i)])


def addAndDelData(myData, standData):
    """
    新增或删除列表内的字典,需增加"addType“字段  add为增加, del 为删除
    @param myData: 自己的数据
    @param standData: 需要替换的数据
    @return:
    """
    assert isinstance(myData, list), "传入数据错误"
    for mydict in myData:
        if mydict.get("addType") == "add":
            mydict.pop("addType")
            standData.insert(0, mydict)
        elif mydict.get("addType") == "del":
            mydict.pop("addType")
            for sData in standData:
                if mydict.get("dimensionType") == sData.get("dimensionType") \
                        and mydict.get("dimension") == sData.get("dimension"):
                    delIndex = standData.index(sData)
                    del standData[delIndex]
                elif mydict.get("permissionType") == sData.get("permissionType") \
                        and mydict.get("documentType") == sData.get("documentType"):
                    delIndex = standData.index(sData)
                    del standData[delIndex]


def selectKeyAndValue(data, **kwargs):
    """
    从返回的数据中,查找value值唯一的字典
    :param data:
    :param kwargs:
                asserInfo = selectKeyAndValue(
                    data=res,
                    name="采购订单",
                )
    :return:
    """
    returnData = None
    if isinstance(data, dict):
        for Key, Value in data.items():
            if isinstance(Value, list):
                returnData = selectKeyAndValue(Value, **kwargs)
                if returnData is None:
                    pass
                else:
                    return returnData

            else:
                for keyKwargs, valueKwargs in zip(kwargs.keys(), kwargs.values()):
                    if keyKwargs == Key and valueKwargs == Value:
                        returnData = data
                        if returnData is None:
                            pass
                        else:
                            return returnData
    elif isinstance(data, list):
        for d in data:
            returnData = selectKeyAndValue(d, **kwargs)
            if returnData is None:
                pass
            else:
                return returnData
    if returnData is None:
        pass
    else:
        return returnData


if __name__ == '__main__':
    # print(getCycleDate('2023-03-29'))
    # print(getCycleDate('2023-03-30'))
    # print(getCycleDate('2023-03-31'))
    # print(zRound(203.8835078534031, 3))
    print(getFileName(filePath={"dddddd"}))
    # print(listDeDuplication(["wlbm003", "WLBM001", "WLBM001", "WLBM001", "WLBM004", "WLBM002"]))
    # print(listDeDuplication([1, 2, 3,4, 3, 2]))
    # price0 = calTaxIncludedPriceWithUnitPrice(unitPrice=176.9911504425, precision=10)
    # price1 = calTaxIncludedPriceWithUnitPrice(unitPrice=176.9911504425, precision=10)
    # price2 = calUnitPriceWithTaxIncludedPrice(taxIncludedPrice=200, precision=10)
    # price3 = calUnitPriceWithTaxIncludedPrice(taxIncludedPrice=200, precision=10)
    pass
    # print(getPrecision(465462.))
    # print(date_time(timezone=None, days=1, hours=0, minutes=0, fmt="%Y-%m-%d"))
    # changePath(midDir="srm-itf-test", rightPath='')
    # print(changePath("SrmInterfaceTest","/resource/files/contract/contract_upload_test.doc"))
    # print(strToDate(dateStr="2023-07-21 00:00:00", days=1, seconds=2, inFmt="%Y-%m-%d %H:%M:%S", outFmt="%Y-%m-%d %H:%M:%S"))
    # print(cleanNoneFromDict({"a": "ccc", "c": None, "dd":None}))
    # listAC = [
    #     {"a": 3.15, "c": None, "dd": None},
    #     {"a": 3.655, "c": None, "dd": None},
    #     {"a": 3.555, "c": None, "dd": None}
    # ]
    # print(sortDictIntInList(listA=listAC, key="a"))
